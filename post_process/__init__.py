import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import ClassVar
from datetime import datetime

import dotenv
import openpyxl
import pandas
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from helper import google_api_helper, api_scopes, rangeconvert

dotenv.load_dotenv()

table_search_range = os.environ["TABLE_SEARCH_RANGE"]

update_sheet_id = os.environ["SCHEDULE_SHEET_ID"]

# TODO: 2022/12/28 これは定数的にしてもいいかな。
msm_anken_number_pattern = re.compile(r"(MA-\d{4}).*")

range_addr_pattern = re.compile(
    r"^(?P<sheetname>.*)!(?P<firstcolumn>[A-Z]+)(?P<firstrow>\d+)"
)

google_cred: Credentials = google_api_helper.get_cledential(
    api_scopes.GOOGLE_API_SCOPES
)


@dataclass
class RenrakukoumokuInfo:
    renrakukoumoku_path: Path
    anken_base_number: str = field(init=False)
    kokyaku_name: str = field(init=False)
    enduser_name: str = field(init=False)
    format_version: str = field(init=False)

    renrakukoumoku_version_pattern: ClassVar[dict] = {
        "before202211": {"kokyaku_name_celaddr": "D6", "enduser_name_celaddr": "D9"},
        "new": {"kokyaku_name_celaddr": "D6", "enduser_name_celaddr": "D10"},
    }

    def __post_init__(self):
        # 念のために前処理でスペースがあれば除去してる
        self.anken_base_number = msm_anken_number_pattern.search(
            self.renrakukoumoku_path.name.replace(" ", "")
        ).group(1)

        renrakukoumoku_ws = openpyxl.load_workbook(self.renrakukoumoku_path).active
        version_pattern_check = renrakukoumoku_ws["B9"].value
        version_cell_addr = {}
        match version_pattern_check:
            case "エンドユーザー":
                version_cell_addr = self.renrakukoumoku_version_pattern["before202211"]
                self.format_version = "before202211"
            case "顧客連絡先":
                version_cell_addr = self.renrakukoumoku_version_pattern["new"]
                self.format_version = "new"
            case _:
                print("連絡項目の内容が不正の可能性があります")
                return None

        self.kokyaku_name = renrakukoumoku_ws[
            version_cell_addr["kokyaku_name_celaddr"]
        ].value
        self.enduser_name = renrakukoumoku_ws[
            version_cell_addr["enduser_name_celaddr"]
        ].value


@dataclass
class CsvFileInfo:
    csv_filepath: Path
    anken_number: str = field(init=False)
    anken_base_number: str = field(init=False)
    hose_type: str = field(init=False)
    hose_attachment_types: set[str] = field(init=False)

    def __post_init__(self):
        # 前処理: ファイル名にスペースとアンダーバーが入ることがあるので、置き換え
        self.anken_number = self.csv_filepath.stem.replace(" ", "").replace("_", "-")

        # 正規表現で MA-0000, MA-0000-1という
        self.anken_base_number = msm_anken_number_pattern.search(
            self.anken_number
        ).group(1)

        csv_pd = pandas.read_csv(self.csv_filepath, encoding="shift-jis")
        hose_parts_pd = csv_pd[csv_pd["品名"] == "ホース(継手付)"]

        self.hose_type = hose_parts_pd["備考"].drop_duplicates().values[0]

        # 型式寸法から、SS,SL,LLを取り出して、重複を外して種類を確定させる
        self.hose_attachment_types = "/".join(
            {item.split("-")[1] for item in hose_parts_pd["型式・寸法"].values}
        )


# TODO:2023-01-12 ここではまだ一括で登録をする作業はできないので、gsheet利用優先で実装中
@dataclass
class EstimateCalcSheetInfo:

    # gsheet_url | openpyxl.ws を受け取るような仕様にする。
    calcsheet_source: str | Path
    anken_number: str = field(init=False)
    anken_base_number: str = field(init=False)

    nouki: str = field(init=False)
    price: str = field(init=False)

    def __post_init__(self):
        # sourceの種類で取り込む処理を変える
        match self.calcsheet_source:
            # excel形式: Pathを指定する
            case Path():
                if self.calcsheet_source.suffix == ".xlsx":

                    pass
                pass
            # gsheet形式: IDの羅列なのでIDが利用できるかはAPIに問い合わせる
            case str():
                try:
                    sheet_service = build("sheets", "v4", credentials=google_cred)

                    # スプレッドシート名を収集して、anken_numberを生成

                    estimate_calc_gsheet_res = (
                        sheet_service.spreadsheets()
                        .get(
                            spreadsheetId=self.calcsheet_source,
                        )
                        .execute()
                    )

                    self.anken_number = msm_anken_number_pattern.search(
                        estimate_calc_gsheet_res["properties"]["title"]
                    ).group(0)
                    self.anken_base_number = msm_anken_number_pattern.search(
                        self.anken_number
                    ).group(1)

                    # 古い仕様のシートと新しいシートの違いで、range_mapを変える。計算結果シートがある場合はそこを参照
                    estimate_calc_sheetnames = [
                        sheets["properties"]["title"]
                        for sheets in estimate_calc_gsheet_res["sheets"]
                    ]
                    # print(estimate_calc_sheetnames)
                    # TODO:2023-01-12 この部分は判断方法と結果を返す関数にした方がいいと思われる。クラスの裏に分離したほうがいいな
                    range_map = {
                        "Sheet1!F17": "price",
                        "Sheet1!F1": "nouki",
                    }

                    if "計算結果" in estimate_calc_sheetnames:
                        range_map = {
                            "'計算結果'!B5": "price",
                            "'計算結果'!B6": "nouki",
                        }

                    # シートの情報を収集して各フィールドへ追加す各
                    range_names = list(range_map.keys())
                    estimate_calc_gsheet_values_res = (
                        sheet_service.spreadsheets()
                        .values()
                        .batchGet(
                            spreadsheetId=self.calcsheet_source,
                            ranges=range_names,
                            valueRenderOption="UNFORMATTED_VALUE",
                        )
                        .execute()
                    )
                    for res_value in estimate_calc_gsheet_values_res.get("valueRanges"):
                        setattr(
                            self,
                            range_map[res_value.get("range")],
                            str(res_value.get("values")[0][0]),
                        )
                    self.nouki = self.nouki.replace(".", "/")

                except HttpError as error:
                    # TODO:2022-12-09 エラーハンドリングは基本行わずここで落とすこと
                    # TODO: 2022/12/28 このエラーは致命的なのでそのままプログラム自体も終了する
                    print(f"An error occurred: {error}")
                    exit()
            # Path, str以外はエラーとする
            case _:
                raise ValueError(
                    f"This source is cant use class:{self.calcsheet_source}"
                )


# TODO:2022-12-29 案件マップというよりは、コンバーターという名前のほうがしっくりくる
@dataclass
class MsmAnkenMap:

    csvfile_info: CsvFileInfo | None = None
    estimate_calcsheet_info: EstimateCalcSheetInfo | None = None
    anken_number: str = field(init=False)
    anken_base_number: str = field(init=False)
    # 連絡項目と関連するものを入れるフィールド
    renrakukoumoku_info: RenrakukoumokuInfo | None = None

    def __post_init__(self):
        # TODO:2022-12-29 今はcsvfile_infoがあることが前提だが、estimate_sheet_infoのどちらか、または両方ある場合の対応を作ること
        self.anken_number = self.select_anken_ref().anken_number
        self.anken_base_number = self.select_anken_ref().anken_base_number

    def select_anken_ref(self):
        anken_ref_obj = None
        if self.csvfile_info:
            anken_ref_obj = self.csvfile_info
        elif self.estimate_calcsheet_info:
            anken_ref_obj = self.estimate_calcsheet_info
        return anken_ref_obj

    # 書くデータのクラスを取り込む関数（set_***）
    # setしたときにanken_base_numberが同じかをチェック
    def set_renrakukoumoku_info(self, renrakukoumoku_info: RenrakukoumokuInfo):

        if (
            renrakukoumoku_info.anken_base_number
            == self.select_anken_ref().anken_base_number
        ):
            self.renrakukoumoku_info = renrakukoumoku_info


@dataclass
class MsmAnkenMapList:

    msmankenmap_list: list[MsmAnkenMap] = field(default_factory=list)

    # 名寄せ用のマップ
    schedule_sheet_map: ClassVar[dict] = {
        "納期": "estimate_calcsheet_info.nouki",
        "金額(税抜)": "estimate_calcsheet_info.price",
        "ガス本数": "csvfile_info.gas_qty",
        "ホース本数": "csvfile_info.hose_qty",
        "ホースタイプ": "csvfile_info.hose_type",
        "利用したホースの接続継手の種類": "csvfile_info.hose_attachment_types",
        "顧客名": "renrakukoumoku_info.kokyaku_name",
        "エンドユーザー": "renrakukoumoku_info.enduser_name",
    }

    # 更新データの生成
    # CSVor estimatesheetがない場合は処理を行わない（連絡項目だけだと正確な案件番号が出ないので）
    def generate_update_sheet_values(self) -> pandas.DataFrame:
        # 入っているmsmankenmapにcsvファイル/見積書計算のデータクラスがない場合は無視する csvfile_info.anken_numberがない場合は無視
        sheet_table = []
        for msmankeninfo in self.msmankenmap_list:
            if not msmankeninfo.anken_number:
                print("案件番号がありません。CSVファイルか計算表を登録してください")
                continue

            # 各種のデータクラスから値を取り出していく
            # TODO:2022-12-29 各データクラスのどれかがない場合の考慮がまだできてない。
            # ない場合は空白とかでいいと思う

            sheet_row = {"msmankennumber": msmankeninfo.anken_number}

            # dataclass_attr.split(".")[0]):CSV, 連絡項目などの収集infoのオブジェクト
            # dataclass_attr.split(".")[1]):それぞれ収集infoオブジェクトのインスタンス属性
            sheet_row = sheet_row | {
                sheet_column_name: getattr(
                    getattr(msmankeninfo, map_dataclass_attr.split(".")[0]),
                    map_dataclass_attr.split(".")[1],
                    None,
                )
                # sheet_column_name: map_dataclass_attr
                for sheet_column_name, map_dataclass_attr in MsmAnkenMapList.schedule_sheet_map.items()
            }
            sheet_table.append(sheet_row)

        # df変換
        return pandas.DataFrame(sheet_table).set_index(["msmankennumber"]).fillna("")


def get_schedule_table_area(
    search_range: str, google_cred
) -> tuple[str, pandas.DataFrame]:

    # 更新対称のセル範囲から値を取得
    try:
        append_values = [
            [],
        ]
        # appendを使ってテーブルの範囲を取得最終行からテーブルの範囲を取得
        sheet_service = build("sheets", "v4", credentials=google_cred)
        append_end_row = (
            sheet_service.spreadsheets()
            .values()
            .append(
                spreadsheetId=update_sheet_id,
                range=search_range,
                body={"values": append_values},
                valueInputOption="USER_ENTERED",
                insertDataOption="INSERT_ROWS",
            )
        ).execute()

        last_low_range = range_addr_pattern.match(
            append_end_row.get("updates").get("updatedRange"),
        ).group("firstrow")

        # print(append_end_row.get("updates").get("updatedRange"))
        # print(last_low_range)

        schedule_range_matcher = range_addr_pattern.match(search_range)

        schedule_table_range = f"{schedule_range_matcher.group(0)}:Q{last_low_range}"
        # print(schedule_table_range)

        schedule_table_res = (
            sheet_service.spreadsheets()
            .values()
            .get(spreadsheetId=update_sheet_id, range=schedule_table_range)
            .execute()
        )

        # 値が取得できた。
        schedule_table = schedule_table_res.get("values")
        # print(schedule_table)

        # df変換

        result_pd = pandas.DataFrame(schedule_table)
        result_pd = result_pd.set_axis(labels=schedule_table[0], axis=1).drop(
            result_pd.index[[0]]
        )
        result_pd = result_pd.set_index(["図番"], drop=False)
        # print(result_pd)
        # exit()
        return result_pd

    except HttpError as error:
        # このエラーは致命的なのでそのままプログラム自体も終了する
        print(f"An error occurred: {error}")
        exit()


def generate_update_valueranges(
    search_range: str, old_pd: pandas.DataFrame, new_pd: pandas.DataFrame
) -> list[dict]:
    """
    oldとnewの表をみて、old側で更新するべきセルのアドレスと値を収集する
    結果はGoogle Sheet APIで受け取れるValueRangeとする
    # https://developers.google.com/sheets/api/reference/rest/v4/spreadsheets.values#ValueRange

    search_range ="'sheet1'!A5:Q5"
    [{"rangeaddr": "value"}, ...]
    """

    # TODO: 2023/01/16 今回はcolumnについては何も考慮していない。A1表記のAを変換して数字にした上で、次のテーブルのループで対象のセルアドレスを生成する必要あり。今回はA列から取ってるからやらなくていいけど、実装の余裕ができたらやろう
    schedule_range_matcher = range_addr_pattern.match(search_range)

    start_row = int(schedule_range_matcher.group("firstrow")) + 1
    # print(f"{schedule_range_matcher} -> ({start_row})")

    result_rangevalues = []
    new_pd_columns = new_pd.columns.tolist()
    new_pd_indexs = new_pd.index.tolist()
    # print(new_pd_columns)

    for table_row_index, (old_index, row) in enumerate(old_pd.iterrows()):
        for table_column_index, (old_column_name, column_by_row) in enumerate(
            row.items()
        ):
            # oldのインデックスとカラム名がnew側にない場合は無視
            if not (old_index in new_pd_indexs and old_column_name in new_pd_columns):
                continue
            # 上書き防止条件: newに値がある, old側に値がない場合は追加する
            if new_pd.loc[old_index, old_column_name] and not column_by_row:
                cell_addr = rangeconvert.rowcol_to_a1(
                    start_row + table_row_index, table_column_index + 1
                )
                # print(new_pd.loc[old_index, old_column_name])
                # 上のrangeaddr, valueの辞書を作る。
                result_rangevalues.append(
                    {
                        "range": cell_addr,
                        "values": [[new_pd.loc[old_index, old_column_name]]],
                    }
                )
    return result_rangevalues


def update_schedule_sheet(update_data: list[dict], google_cred):
    # 最後にbatchUpdateを行うためのデータ構造を生成して流し込む
    update_value_body = {
        "data": [update_data],
        "valueInputOption": "USER_ENTERED",
    }

    sheet_service = build("sheets", "v4", credentials=google_cred)
    try:

        update_gsheet_res = (
            sheet_service.spreadsheets()
            .values()
            .batchUpdate(spreadsheetId=update_sheet_id, body=update_value_body)
            .execute()
        )
        return update_gsheet_res
    except HttpError as error:
        # TODO:2022-12-09 エラーハンドリングは基本行わずここで落とすこと
        print(f"An error occurred: {error}")
        exit()