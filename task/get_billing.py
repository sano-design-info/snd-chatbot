import json
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from dateutil import parser
from dateutil.relativedelta import relativedelta
from googleapiclient.discovery import build
from openpyxl.styles import Border, Side
from zoneinfo import ZoneInfo
from api.googleapi import sheet_data_mapper

import chat.card
from api import googleapi

from helper import EXPORTDIR_PATH, ROOTDIR, chatcard, load_config
from helper.regexpatterns import BILLING_DURARION, MSM_ANKEN_NUMBER
from task import BaseTask, ProcessData

# `2020-01-01` のフォーマットのみ受け付ける
START_DATE_FORMAT = "%Y-%m-%d"


config = load_config.CONFIG
# MISUMI_TORIHIKISAKI_ID = config.get("mfci").get("TORIHIKISAKI_ID")
# QUOTE_PER_PAGE = config.get("mfci").get("QUOTE_PER_PAGE")

SCRIPT_CONFIG = config.get("get_billing")
GENERATE_QUOTES_CONFIG = config.get("generate_quotes")
TORIHIKISAKI_NAME = config.get("general").get("TORIHIKISAKI_NAME")

# 請求書生成:

# 見積書のファイル一覧を記録するGoogleスプレッドシートのID
QUOTE_FILE_LIST_GSHEET_ID = GENERATE_QUOTES_CONFIG.get("QUOTE_FILE_LIST_GSHEET_ID")
# テンプレートに入力するセルマッピング・JSONファイルのパス
QUOTE_TEMPLATE_CELL_MAPPING_JSON_PATH = GENERATE_QUOTES_CONFIG.get(
    "QUOTE_TEMPLATE_CELL_MAPPING_JSON_PATH"
)
with open(QUOTE_TEMPLATE_CELL_MAPPING_JSON_PATH, "r", encoding="utf-8") as f:
    quote_template_cell_mapping_dict = json.load(f)


# 請求書のファイル一覧を記録するGoogleスプレッドシートのID
# https://docs.google.com/spreadsheets/d/1_x1yBpm34FWJ1shXQeUZVwwqp8CdT34t9zH5T-urzHs
INVOICE_FILE_LIST_GSHEET_ID = SCRIPT_CONFIG.get("INVOICE_FILE_LIST_GSHEET_ID")
# GoogleスプレッドシートのテンプレートID
# https://docs.google.com/spreadsheets/d/1444u_Gu9-1VI2C5EmnX5uE3h-QwrTrroGpQzZNupzNI
INVOICE_TEMPLATE_GSHEET_ID = SCRIPT_CONFIG.get("INVOICE_TEMPLATE_GSHEET_ID")
# テンプレートに入力するセルマッピング・JSONファイルのパス
INVOICE_TEMPLATE_CELL_MAPPING_JSON_PATH = SCRIPT_CONFIG.get(
    "INVOICE_TEMPLATE_CELL_MAPPING_JSON_PATH"
)
# 請求書ののGoogleスプレッドシート保存先
# https://drive.google.com/drive/folders/1fM40M8T5Yhj16nfpLQ1i_oPcd0iJqFHt
INVOICE_GSHEET_SAVE_DIR_IDS = SCRIPT_CONFIG.get("INVOICE_GSHEET_SAVE_DIR_IDS")
# 請求書のPDFと請求一覧のExcel保存先
# https://drive.google.com/drive/folders/1kYA0eGhNUhYPjkWiGRP018AR6yPs2PGD
INVOICE_DOC_SAVE_DIR_IDS = SCRIPT_CONFIG.get("INVOICE_DOC_SAVE_DIR_IDS")

# 各定数から全体で使う変数を作成
export_billing_dirpath = EXPORTDIR_PATH / "billing"
export_billing_dirpath.mkdir(parents=True, exist_ok=True)

with open(INVOICE_TEMPLATE_CELL_MAPPING_JSON_PATH, "r", encoding="utf-8") as f:
    invoice_template_cell_mapping_dict = json.load(f)

# 本日の日付を全体で使うためにここで宣言
today_datetime = datetime.now(ZoneInfo("Asia/Tokyo"))
BILLING_PDFFILEPATH = (
    export_billing_dirpath / f"{today_datetime:%Y%m}_ミスミ配管請求書.pdf"
)
BILLING_LIST_EXCELPATH = (
    export_billing_dirpath / f"{today_datetime:%Y%m}_ミスミ配管納品一覧.xlsx"
)

# API Session
# TODO:2023-09-19 ここはtry exceptで囲む
google_cred = googleapi.get_cledential(googleapi.API_SCOPES)
gsheet_service = build("sheets", "v4", credentials=google_cred)
gmail_service = build("gmail", "v1", credentials=google_cred)
gdrive_service = build("drive", "v3", credentials=google_cred)


# mfcloudのセッション作成
# mfcl_session = MFCIClient().get_session()

# チャット用の認証情報を取得
google_sa_cred = googleapi.get_cledential_by_serviceaccount(googleapi.CHAT_API_SCOPES)
chat_service = build("chat", "v1", credentials=google_sa_cred)

spacename = config.get("google").get("CHAT_SPACENAME")
bot_header = chatcard.bot_header


def generate_dl_numbers(dl_numbers_str: str) -> list:
    """
    変換したい見積書の番号を指定したときに端的な番号にする
    例:1,2,3-5,7 => 1,2,3,4,5,7

    """
    dl_numbers = list()
    for dl_number_s in dl_numbers_str.split(","):
        if dl_number_s.isdigit():
            dl_numbers.append(int(dl_number_s))
        elif "-" in dl_numbers_str:
            dl_number_range = dl_number_s.split("-")
            now_number = int(dl_number_range[0])
            end_number = int(dl_number_range[1])
            while now_number <= end_number:
                dl_numbers.append(now_number)
                now_number += 1
    return dl_numbers


@dataclass
class QuoteData:
    """
    見積情報を簡素にまとめたデータ構造

    args:
        durarion_src: 見積書にある品目の詳細
        price: 見積書の合計金額
        hinmoku_title: 見積書にある品目のタイトル

    >>> qdd = {"durarion_src":"納期 10/20", "price":10000, "hinmoku_title":"MA-9901 ガススプリング配管図"}
    >>> qd = QuoteData(**qdd)

    """

    # durarionは最初にstrとして取り込んで、クラス生成時に正規表現で綺麗にする
    durarion_src: str = ""
    price: float = 0
    hinmoku_title: str = ""

    only_katasiki: str = field(init=False)
    durarion: str = field(init=False)

    def __post_init__(self):
        self.only_katasiki = MSM_ANKEN_NUMBER.match(self.hinmoku_title).group(0)
        self.durarion = BILLING_DURARION.match(self.durarion_src).group("durarion")


@dataclass
class BillingInfo:
    """作成する請求書情報のデータ構造"""

    price: float
    billing_title: str
    hinmoku_title: str


def convert_dict_to_gsheet_tamplate(
    invoice_number: str, invoice_title: str, hinmoku_name: str, hinmoku_price: float
) -> dict:
    # 請求書へ書き込むデータを作る
    # ここで作成するデータは、INVOICE_TEMPLATE_CELL_MAPPING_JSON_PATHのテンプレートに合わせたデータを作成する

    hinmoku = {
        "name": hinmoku_name,
        "detail": "",
        "price": int(hinmoku_price),
        "quantity": 1,
        "zeiritu": "10%",
    }

    # TODO:2024-02-05 これは設定にする
    invoice_id_prefix = "DCCF6E"
    # 請求書番号を生成
    invoice_id = f"{invoice_id_prefix}-I-{invoice_number}"

    today_datetime = datetime.now()
    return {
        "customer_name": TORIHIKISAKI_NAME,
        "invoice_id": invoice_id,
        "title": invoice_title,
        # 日付は実行時の日付を利用
        "invoice_date": today_datetime.strftime(START_DATE_FORMAT),
        "due_date": "",
        "note": "詳細は別添付の明細をご確認ください",
        "item_table": [hinmoku],
    }


def set_border_style(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    num_rows: int,
    column_nums: list[int],
    start_num_row: int = 1,
):
    """
    ワークシートに罫線を入れる。
    """

    # 罫線の設定
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # B6からD6までのセルに上下左右に罫線をいれる
    for row in range(start_num_row, num_rows + start_num_row):
        for col in column_nums:
            cell = ws.cell(row=row, column=col)
            cell.border = border
            # D6の表示形式を日本円の通貨表記にする
            if col == 4:
                cell.number_format = "[$¥-ja-JP]#,##0;-[$¥-ja-JP]#,##0"


# TODO:2024-02-12 ここはgoogleスプレッドシートに保存して、excelのファイルとしてダウンロードさせる
def generate_invoice_list_excel(
    billing_target_quotes: list[QuoteData],
) -> Path:
    """
    請求対象一覧をxlsxファイルに出力する
    """
    # テンプレートの形式に沿った数字の設定
    row_start = 6
    # BillingTargetQuoteの構造に従ったマップ
    column_fieldmap = {
        "only_katasiki": 2,
        "durarion": 3,
        "price": 4,
    }

    wb = openpyxl.load_workbook(
        str((ROOTDIR / "templates/billing_list_template.xlsx"))
    )
    ws = wb.active

    # A1セルに"2023年**月請求一覧"と入れる。**は今月の月
    ws["B1"] = f"{today_datetime:%Y年%m月}請求一覧"
    # D1セルには今日の日付を入れる。フォーマットは"2023年4月26日" f-stringで入れる。
    ws["D1"] = f"{today_datetime:%Y年%m月%d日}"

    # 6行目からデータを入れる。
    for row, quote in enumerate(billing_target_quotes, start=row_start):
        for key, col in column_fieldmap.items():
            ws.cell(row=row, column=col).value = getattr(quote, key)

    # 罫線を入れる。set_border_styleを使う
    set_border_style(
        ws, len(billing_target_quotes), column_fieldmap.values(), row_start
    )

    # ファイルを保存する
    # TODO:2023-08-29 ここのファイル名の戻り値って意味ある？
    wb.save(BILLING_LIST_EXCELPATH)
    return BILLING_LIST_EXCELPATH


def str_to_datetime_with_dateutil(date_str):
    """
    日付の文字列 月/日 をdatetimeオブジェクトに変換する。
    現在の年を考慮して、今年の12月に1月または2月が入っていた場合、来年の1月または2月にする。

    注意:この考慮は少し曖昧で、12月だから1月2月が来年というよりは、現在の月より前の月だから来年という考え方が正しいはず。
    12月頃に通常発生する自然な表現として採用している。11月に1月2月が入っていたら、といったたらればは今は考えない。

    args:
        date_str: 日付の文字列 月/日
    return:
        datetimeオブジェクト
    """
    # 現在の年を取得
    current_year = datetime.now().year

    # 現在の月を取得
    current_month = datetime.now().month

    # dateutilを使用して日付文字列からdatetimeオブジェクトを生成
    # date = datetime.strptime(date_str, "%m/%d")
    date = parser.parse(date_str)

    # 現在の月が12月で、生成された日付の月が1月または2月の場合、年を+1する
    if current_month == 12 and (date.month == 1 or date.month == 2):
        date = date.replace(year=current_year + 1)

    # タイムゾーンを入れて返す
    return date


# TODO:2024-02-12 test_get_billingでテストを書くこと
def is_range_date(
    target_date: datetime, start_date: datetime, end_date: datetime | None = None
) -> bool:
    """
    指定した日付が、指定した日付の範囲内にあるかどうかを判定する
    beforeが一つのみ指定されている場合は、beforeがafterより前になっていたらエラーを出す
    args:
        target_date: 判定する日付
        start_date: 範囲の開始日
        end_date: 範囲の終了日。オプショナル。指定しない場合はstart_dateのみで判定

    return:
        範囲内にあればTrue、範囲外ならFalse
    """

    # startのみ指定がある場合、それだけで比較結果を返す
    if end_date is None:
        return target_date >= start_date

    # startよりendが前の場合はエラーを出す
    if start_date > end_date:
        raise ValueError("開始日は終了日より前でなければなりません。")

    # start,endがあり、start,endの範囲が適切な場合、範囲内かどうかを判断
    return start_date <= target_date <= end_date


# 請求書の金額合計のデータを生成する
def generate_invoice_data(
    quote_checked_list: list[QuoteData],
) -> BillingInfo:
    """
    請求書にする見積書の金額を合計する。
    表示用に件数とBillingInfoのタプルを返す

    args:
        quote_checked_list: 請求書にする見積書のリスト
    return:
        件数とBillingInfoのタプル

    """
    return BillingInfo(
        sum((i.price for i in quote_checked_list)),
        "ガススプリング配管図作製費",
        f"{today_datetime:%Y年%m月}請求分",
    )


# メール下書きを作成する
def set_draft_mail(attchment_filepaths: list[Path]) -> dict:
    """
    タイトルと本文を入力してメール下書きを作成する
    タイトルの例 "2023年03月請求書送付について"
    """
    # タイトルは日付が入ったもの。例:2023年03月請求書送付について
    mailtitle = SCRIPT_CONFIG.get("mail_template_title").replace(
        "{{datetime}}", f"{today_datetime:%Y年%m月}"
    )

    mailbody = SCRIPT_CONFIG.get("mail_template_body")
    mailto = SCRIPT_CONFIG.get("mail_to")
    mailcc = SCRIPT_CONFIG.get("mail_cc", "")

    # メール下書きを作成する
    return googleapi.append_draft(
        gmail_service, mailto, mailcc, mailtitle, mailbody, attchment_filepaths
    )


def get_quote_gsheet_by_quote_list_gsheet(
    gsheet_service, per_page_length: int
) -> list | None:
    # Googleスプレッドシートの見積書一覧をみて、最後尾から必要な件数の見積書スプレッドシートのURLリストを取得する

    # 見積計算表の生成元シートの最大行範囲を取得
    range_name = "見積書管理!C2:C"
    result = (
        gsheet_service.spreadsheets()
        .values()
        .get(spreadsheetId=QUOTE_FILE_LIST_GSHEET_ID, range=range_name)
        .execute()
    )
    values = result.get("values", [])

    if not values:
        return None
    else:
        # 最後の行を取得
        last_row = len(values)
        print(f"{range_name} 最後の行: {last_row}")
        return values[last_row - per_page_length : last_row]


def get_values_by_range(
    gsheet_service, spreadsheet_id, name_and_range_dict: dict
) -> dict:
    """{"意味名":"セル番号"}の辞書をもとに、Googleスプレッドシートの値を取得する。バッチで複数getする
    ここでは、シートの値は必ず1行1列の値として取得することを前提としている。
    戻り値は、{"意味名": "セルの値"}の辞書
    """
    result = (
        gsheet_service.spreadsheets()
        .values()
        .batchGet(
            spreadsheetId=spreadsheet_id,
            ranges=list(name_and_range_dict.values()),
            majorDimension="COLUMNS",
        )
        .execute()
    )
    # 取得した値を辞書にして返す
    return {
        key: value.get("values",[[""]])[0][0]
        for key, value in zip(name_and_range_dict, result.get("valueRanges"))
    }


def get_hinmoku_celladdrs_by_gsheet():
    # 見積スプレッドシートURLから 見積書の品目名と納期と金額を取得。品目は最初の1行のみ取得。
    # 情報のセルアドレスはテンプレートのセルマッピングに従う。
    hinmoku_table_celladdr_column = (
        quote_template_cell_mapping_dict.get("tables").get("item_table").get("columns")
    )
    hinmoku_table_startrow = (
        quote_template_cell_mapping_dict.get("tables").get("item_table").get("startRow")
    )
    quote_single_celladdr = quote_template_cell_mapping_dict.get("singlecell")
    return {
        "hinmoku_name": f"{hinmoku_table_celladdr_column.get('name')}{hinmoku_table_startrow}",
        "hinmoku_detail": f"{hinmoku_table_celladdr_column.get('detail')}{hinmoku_table_startrow}",
        "hinmoku_price": f"{hinmoku_table_celladdr_column.get('price')}{hinmoku_table_startrow}",
        "quote_date": quote_single_celladdr.get("quote_date"),
        "quote_id": quote_single_celladdr.get("quote_id"),
    }


class PrepareTask(BaseTask):
    def execute_task(self) -> list[tuple[QuoteData, bool]]:
        # 見積書一覧を取得

        # TODO: 2024-02-06 この100件を取り出す方法は本来は動作が微妙
        # 見積書の納期日をもとに40日前までの見積書を取得するようにするほうがベスト。
        # もしくはスケジュール表の実納期日をもとにすると良さそう

        # 見積管理表の最後尾100件の見積書スプレッドシートのURLリストを取得する
        # URL（https://docs.google.com/spreadsheets/d/fasfdasfa_IDS_fsdfadsfa）からID(fasfdasfa_IDS_fsdfadsfa)を取得する
        quote_gsheet_url_list_under_100 = get_quote_gsheet_by_quote_list_gsheet(
            gsheet_service, 100
        )
        print(f"見積書のURLリスト: {quote_gsheet_url_list_under_100}")

        quote_gsheet_id_list_under_100 = [
            i[0].split("/")[-1] for i in quote_gsheet_url_list_under_100
        ]
        print(f"見積書のIDリスト: {quote_gsheet_id_list_under_100}")

        # URLリストからAPIで見積書の情報を取得。GoogleスプレッドシートのIDから情報を取得する。
        quota_values_list_extracted_from_gsheet = [
            get_values_by_range(gsheet_service, id, get_hinmoku_celladdrs_by_gsheet())
            for id in quote_gsheet_id_list_under_100
        ]

        print(f"見積書の情報: {quota_values_list_extracted_from_gsheet}")

        # 見積一覧から必要情報を収集
        # 見積作成時から40日前まで and 品目のフォーマットがミスミの案件番号（正規表現で判断）かでフィルター
        from_date = datetime.now(ZoneInfo("Asia/Tokyo")) + timedelta(days=-40)
        quote_data_list = [
            QuoteData(
                durarion_src=quote_values["hinmoku_detail"],
                price=float(quote_values["hinmoku_price"]),
                hinmoku_title=quote_values["hinmoku_name"],
            )
            for quote_values in quota_values_list_extracted_from_gsheet
            if is_range_date(
                datetime.strptime(quote_values["quote_date"], "%Y/%m/%d").replace(
                    tzinfo=ZoneInfo("Asia/Tokyo")
                ),
                from_date,
            ) and MSM_ANKEN_NUMBER.match(quote_values["hinmoku_name"])
        ]

        # デフォルト表示の選択マーク用のリストを作成
        # 納期（duration）を使って複数選択のデフォルト選択をマーク
        # 期限設定: 実行日（今日）から1か月前の日付をマーク。実行日は月末想定なので、今月末を計算する
        
        # 見積書の一覧を表示して、選択させる
        return [
            (
                quote_data,
                is_range_date(
                    str_to_datetime_with_dateutil(quote_data.durarion).replace(
                        tzinfo=ZoneInfo("Asia/Tokyo")
                    ), 
                    # 実行日の月初め
                    today_datetime.replace(day=1),
                    # ここは今月末を取得する処理に変更する
                    today_datetime.replace(day=1) + relativedelta(months=1) - timedelta(days=1)

                ),
            )
            for quote_data in quote_data_list
        ]

    def execute_task_by_chat(self):
        result = self.execute_task()
        # 選択時のチェックボックス
        quotelist_checkbox = chat.card.genwidget_checkboxlist(
            "請求書にする見積書を選択してください。",
            "quoteitems",
            [
                chat.card.SelectionInputItem(
                    f"{quotedata.only_katasiki} | {quotedata.price} |{quotedata.durarion}",
                    json.dumps(asdict(quotedata)),
                    selected=checkbool,
                )
                for quotedata, checkbool in result
            ],
        )
        # 設定カードを生成
        config_body = chat.card.create_card(
            "config_card__get_billing",
            header=bot_header,
            widgets=[
                quotelist_checkbox,
                # ボタンを追加
                chat.card.genwidget_buttonlist(
                    [
                        chat.card.gencomponent_button(
                            "タスク実行確認", "confirm__get_billing"
                        ),
                        chat.card.gencomponent_button("キャンセル", "cancel_task"),
                    ]
                ),
            ],
        )
        return googleapi.create_chat_message(chat_service, spacename, config_body)


class MainTask(BaseTask):
    def execute_task(self, process_data: ProcessData | None = None) -> dict:
        ask_choiced_quote_list = process_data["task_data"].get("choiced_quote_list")
        invoice_data = generate_invoice_data(ask_choiced_quote_list)

        # TODO:2024-02-12 ここの変更も必須
        # * 見積一覧をexcelで記入する -> Googleスプレッドシート化してダウンロードできたら行う

        generate_invoice_list_excel(ask_choiced_quote_list)
        # excelファイルをGoogleドライブへ保存
        upload_xlsx_result = googleapi.upload_file(
            gdrive_service,
            BILLING_LIST_EXCELPATH,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            INVOICE_DOC_SAVE_DIR_IDS,
        )
        print(
            f"請求分一覧xlsxファイルをGoogleドライブへ保存しました。: {upload_xlsx_result.get('id')}"
        )

        # * 請求書管理表の最後尾の請求書番号を取得
        updated_invoice_manage_gsheet = googleapi.append_sheet(
            gsheet_service,
            INVOICE_FILE_LIST_GSHEET_ID,
            "請求書管理",
            [['=TEXT(ROW()-1,"0000")', "", "", ""]],
            "USER_ENTERED",
            "INSERT_ROWS",
            True,
        )
        # 請求書番号を取得
        invoice_number = (
            updated_invoice_manage_gsheet.get("updates")
            .get("updatedData")
            .get("values")[0][0]
        )
        print(f"請求書の管理表から番号を生成しました。: {invoice_number}")

        # * 請求書の作成
        converted_invoice_dict = convert_dict_to_gsheet_tamplate(
            invoice_number,
            invoice_data.billing_title,
            invoice_data.hinmoku_title,
            invoice_data.price,
        )

        # googleスプレッドシートの見積書テンプレートを複製する
        invoice_file_id = googleapi.dupulicate_file(
            gdrive_service,
            INVOICE_TEMPLATE_GSHEET_ID,
            BILLING_PDFFILEPATH.stem,
        )

        # 請求書スプレッドシートのファイル名と保存先を設定
        _ = googleapi.update_file(
            gdrive_service,
            file_id=invoice_file_id,
            body=None,
            add_parents=INVOICE_GSHEET_SAVE_DIR_IDS,
            fields="id, parents",
        )

        # 請求書スプレッドシートへ請求情報を記入
        sheet_data_mapper.write_data_to_sheet(
            gsheet_service,
            invoice_file_id,
            converted_invoice_dict,
            invoice_template_cell_mapping_dict,
        )

        # 請求書のPDFをダウンロード
        googleapi.export_pdf_by_driveexporturl(
            google_cred.token,
            invoice_file_id,
            BILLING_PDFFILEPATH,
            {
                "gid": "0",
                "size": "7",
                "portrait": "true",
                "fitw": "true",
                "gridlines": "false",
            },
        )

        # 請求書のPDFをGoogleドライブへ保存
        upload_pdf_result = googleapi.upload_file(
            gdrive_service,
            BILLING_PDFFILEPATH,
            "application/pdf",
            "application/pdf",
            INVOICE_DOC_SAVE_DIR_IDS,
        )

        print(
            f"請求書のPDFをGoogleドライブへ保存しました。: {upload_pdf_result.get('id')}"
        )

        # 請求書のGoogleスプレッドシートとPDFのURLを請求書管理表に記録
        _ = googleapi.update_sheet(
            gsheet_service,
            INVOICE_FILE_LIST_GSHEET_ID,
            # 請求書管理表の番号のセルアドレスを元にB列から追加する
            updated_invoice_manage_gsheet.get("updates")
            .get("updatedRange")
            .replace("A", "B"),
            [
                [
                    BILLING_PDFFILEPATH.stem,
                    f"http://docs.google.com/spreadsheets/d/{invoice_file_id}",
                    f"http://drive.google.com/file/d/{upload_pdf_result.get('id')}",
                ]
            ],
        )

        print("一覧と請求書生成しました")
        print(
            f"一覧xlsxファイルパス:{BILLING_LIST_EXCELPATH}\n請求書pdf:{BILLING_PDFFILEPATH}"
        )

        return set_draft_mail([BILLING_LIST_EXCELPATH, BILLING_PDFFILEPATH])

    # チャット用のタスクメソッド
    def execute_task_by_chat(
        self, process_data: ProcessData | None = None
    ) -> dict | str:
        result = self.execute_task(process_data)
        # チャット用のメッセージを作成する
        send_message_body = chat.card.create_card(
            "result_card__get_billing",
            header=bot_header,
            widgets=[
                chat.card.genwidget_textparagraph(
                    f"請求書と請求一覧を作成しました。: {result.get('id')}"
                ),
            ],
        )
        # send_message_body.update({"actionResponse": {"type": "NEW_MESSAGE"}})
        return googleapi.create_chat_message(chat_service, spacename, send_message_body)
