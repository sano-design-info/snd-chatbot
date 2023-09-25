# coding: utf-8
import json
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import questionary
from dateutil.relativedelta import relativedelta
from dateutil import parser
import openpyxl
from googleapiclient.discovery import build
from openpyxl.styles import Border, Side
from api import googleapi

from helper import load_config, EXPORTDIR_PATH
from api.mfcloud_api import (
    MFCIClient,
    download_billing_pdf,
    get_quotes,
    create_item,
    create_invoice_template_billing,
    attach_billingitem_into_billing,
)
from helper.regexpatterns import BILLING_DURARION, MSM_ANKEN_NUMBER

# `2020-01-01` のフォーマットのみ受け付ける
START_DATE_FORMAT = "%Y-%m-%d"


config = load_config.CONFIG
MISUMI_TORIHIKISAKI_ID = config.get("mfci").get("TORIHIKISAKI_ID")
QUOTE_PER_PAGE = config.get("mfci").get("QUOTE_PER_PAGE")
SCRIPT_CONFIG = config.get("get_billing")

export_billing_dirpath = EXPORTDIR_PATH / "billing"
export_billing_dirpath.mkdir(parents=True, exist_ok=True)

# 本日の日付を全体で使うためにここで宣言
today_datetime = datetime.now()
BILLING_PDFFILEPATH = export_billing_dirpath / f"{today_datetime:%Y%m}_ミスミ配管請求書.pdf"
BILLING_LIST_EXCELPATH = (
    export_billing_dirpath / f"{today_datetime:%Y%m}_ミスミ配管納品一覧.xlsx"
)

# API Session
# TODO:2023-09-19 ここはtry exceptで囲む
google_cred = googleapi.get_cledential(googleapi.API_SCOPES)
gmail_service = build("gmail", "v1", credentials=google_cred)

# mfcloudのセッション作成
mfcl_session = MFCIClient().get_session()


def generate_dl_numbers(dl_numbers_str: str) -> list:
    """
    変換したい見積書の番号を指定したときに端的な番号にする
    例:1,2,3-5,7 => 1,2,3,4,5,7

    """
    dl_numbers = list()
    for dl_number_s in dl_numbers_str.split(","):
        if dl_number_s.isdigit():
            dl_numbers.append(int(dl_number_s))
        elif "-" in dl_numbers_str:
            dl_number_range = dl_number_s.split("-")
            now_number = int(dl_number_range[0])
            end_number = int(dl_number_range[1])
            while now_number <= end_number:
                dl_numbers.append(now_number)
                now_number += 1
    return dl_numbers


@dataclass
class QuoteData:
    """
    見積情報を簡素にまとめたデータ構造

    args:
        durarion_src: 見積書にある品目の詳細
        price: 見積書の合計金額
        hinmoku_title: 見積書にある品目のタイトル
    """

    # durarionは最初にstrとして取り込んで、クラス生成時に正規表現で綺麗にする
    durarion_src: str = ""
    price: float = 0
    hinmoku_title: str = ""

    only_katasiki: str = field(init=False)
    durarion: str = field(init=False)

    def __post_init__(self):
        self.only_katasiki = MSM_ANKEN_NUMBER.match(self.hinmoku_title).group(0)
        self.durarion = BILLING_DURARION.match(self.durarion_src).group("durarion")


@dataclass
class BillingInfo:
    """作成する請求書情報のデータ構造"""

    price: float
    billing_title: str
    hinmoku_title: str


def generate_json_mfci_billing_item(billing_info: BillingInfo) -> dict:
    """
    請求情報を元に、MFクラウド請求書APIで使う請求書書向け品目用のjson文字列を生成する。
    結果は辞書形式で返す。
    """

    item_json_template = """
    {
        "name": "品目",
        "detail": "",
        "unit": "0",
        "price": 0,
        "quantity": 1,
        "excise": "ten_percent"
    }
    """

    # jsonでロードする
    billing_item = json.loads(item_json_template)

    # 結果をjsonで返す
    billing_item["name"] = "{} ガススプリング配管図".format(billing_info.hinmoku_title)
    billing_item["quantity"] = 1
    billing_item["price"] = int(billing_info.price)
    billing_item["name"] = billing_info.hinmoku_title

    return billing_item


def generate_billing_info_json(billing_info: BillingInfo) -> dict:
    """
    MFクラウド請求書APIで使う請求書作成のjsonを生成する
    結果は辞書形式で返す。
    """

    billing_json_template = """
    {
        "department_id": "",
        "billing_date": "2020-05-08",
        "due_date": "2020-05-08",
        "title": "請求書タイトル",
        "tags": "佐野設計自動生成",
        "note": "詳細は別添付の明細をご確認ください",
        "items": [
        ]
    }
    """
    # jsonでロードする
    billing_info_json = json.loads(billing_json_template)

    # department_idはミスミのものを利用
    billing_info_json["department_id"] = MISUMI_TORIHIKISAKI_ID

    # 各情報を入れる
    billing_info_json["title"] = billing_info.billing_title
    billing_info_json["billing_date"] = today_datetime.strftime(START_DATE_FORMAT)

    # due_dateは今月末にする
    # ref: https://zenn.dev/wtkn25/articles/python-relativedelta#%E6%9C%88%E5%88%9D%E3%80%81%E6%9C%88%E6%9C%AB%E3%80%81%E5%85%88%E6%9C%88%E5%88%9D%E3%80%81%E5%85%88%E6%9C%88%E6%9C%AB%E3%80%81%E7%BF%8C%E6%9C%88%E5%88%9D%E3%80%81%E7%BF%8C%E6%9C%88%E6%9C%AB
    billing_info_json["due_date"] = (
        today_datetime + relativedelta(months=+1, day=1, days=-1)
    ).strftime(START_DATE_FORMAT)

    return billing_info_json


# 見積書一覧を元にテンプレの行を生成
def generate_billing_pdf(mfci_session, billing_info: BillingInfo) -> Path:
    """
    MFクラウド請求書APIを使って請求書を生成する。戻り値はpdfファイルのパス
    """

    # 空の請求書作成
    billing_res = create_invoice_template_billing(
        mfci_session, generate_billing_info_json(billing_info)
    )
    # 品目を作成
    billing_item_data = generate_json_mfci_billing_item(billing_info)
    billing_item_res = create_item(mfci_session, billing_item_data)
    # 請求書へ品目を追加する
    attach_billingitem_into_billing(
        mfci_session, billing_res["id"], billing_item_res["id"]
    )

    print("請求書に変換しました")

    # pdfファイルのバイナリをgetする
    download_billing_pdf(mfci_session, billing_res["pdf_url"], BILLING_PDFFILEPATH)

    return BILLING_PDFFILEPATH


def set_border_style(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    num_rows: int,
    column_nums: list[int],
    start_num_row: int = 1,
):
    """
    ワークシートに罫線を入れる。
    """

    # 罫線の設定
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # B6からD6までのセルに上下左右に罫線をいれる
    for row in range(start_num_row, num_rows + start_num_row):
        for col in column_nums:
            cell = ws.cell(row=row, column=col)
            cell.border = border
            # D6の表示形式を日本円の通貨表記にする
            if col == 4:
                cell.number_format = "[$¥-ja-JP]#,##0;-[$¥-ja-JP]#,##0"


def generate_billing_list_excel(
    billing_target_quotes: list[QuoteData],
) -> Path:
    """
    請求対象一覧をxlsxファイルに出力する
    """
    # テンプレートの形式に沿った数字の設定
    row_start = 6
    # BillingTargetQuoteの構造に従ったマップ
    column_fieldmap = {
        "only_katasiki": 2,
        "durarion": 3,
        "price": 4,
    }

    parent_dirpath = Path(__file__).parents[0]

    wb = openpyxl.load_workbook(
        str((parent_dirpath / "itemparser/billing_list_template.xlsx"))
    )
    ws = wb.active

    # A1セルに"2023年**月請求一覧"と入れる。**は今月の月
    ws["B1"] = f"{today_datetime:%Y年%m月}請求一覧"
    # D1セルには今日の日付を入れる。フォーマットは"2023年4月26日" f-stringで入れる。
    ws["D1"] = f"{today_datetime:%Y年%m月%d日}"

    # 6行目からデータを入れる。
    for row, quote in enumerate(billing_target_quotes, start=row_start):
        for key, col in column_fieldmap.items():
            ws.cell(row=row, column=col).value = getattr(quote, key)

    # 罫線を入れる。set_border_styleを使う
    set_border_style(
        ws, len(billing_target_quotes), column_fieldmap.values(), row_start
    )

    # ファイルを保存する
    # TODO:2023-08-29 ここのファイル名の戻り値って意味ある？
    wb.save(BILLING_LIST_EXCELPATH)
    return BILLING_LIST_EXCELPATH


def str_to_datetime_with_dateutil(date_str):
    """
    日付の文字列 月/日 をdatetimeオブジェクトに変換する。
    現在の年を考慮して、今年の12月に1月または2月が入っていた場合、来年の1月または2月にする。

    注意:この考慮は少し曖昧で、12月だから1月2月が来年というよりは、現在の月より前の月だから来年という考え方が正しいはず。
    12月頃に通常発生する自然な表現として採用している。11月に1月2月が入っていたら、といったたらればは今は考えない。

    args:
        date_str: 日付の文字列 月/日
    return:
        datetimeオブジェクト
    """
    # 現在の年を取得
    current_year = datetime.now().year

    # 現在の月を取得
    current_month = datetime.now().month

    # dateutilを使用して日付文字列からdatetimeオブジェクトを生成
    # date = datetime.strptime(date_str, "%m/%d")
    date = parser.parse(date_str)

    # 現在の月が12月で、生成された日付の月が1月または2月の場合、年を+1する
    if current_month == 12 and (date.month == 1 or date.month == 2):
        date = date.replace(year=current_year + 1)

    return date


def is_target_date(t_date: datetime, start_date: datetime, end_date: datetime) -> bool:
    """
    指定した日付が、指定した日付の範囲内にあるかどうかを判定する
    args:
        t_date: 判定したい日付
        start_date: 範囲の開始日:end_dateより手前の日付を検証する
        end_date: 範囲の終了日
    return:
        t_dateがstart_dateとend_dateの間にあるかどうかの真偽値
    """
    # start_dateがend_dateより手前か確認する。違ったらエラー:valueerrorを出す
    if start_date > end_date:
        raise ValueError("start_dateはend_dateより前にしてください")

    return start_date.date() <= t_date.date() <= end_date.date()


# メール下書きを作成する
def set_draft_mail(attchment_filepaths: list[Path]) -> None:
    """
    タイトルと本文を入力してメール下書きを作成する
    タイトルの例 "2023年03月請求書送付について"
    """
    # タイトルは日付が入ったもの。例:2023年03月請求書送付について
    mailtitle = SCRIPT_CONFIG.get("mail_template_title").replace(
        "{{datetime}}", f"{today_datetime:%Y年%m月}"
    )

    mailbody = SCRIPT_CONFIG.get("mail_template_body")
    mailto = SCRIPT_CONFIG.get("mail_to")
    mailcc = SCRIPT_CONFIG.get("mail_cc", "")

    # メール下書きを作成する
    googleapi.append_draft(
        gmail_service, mailto, mailcc, mailtitle, mailbody, attchment_filepaths
    )


def main():
    # 見積書一覧を取得
    quote_result = get_quotes(
        mfcl_session,
        QUOTE_PER_PAGE,
    )

    # 取引先、実行時から40日前まででフィルター
    from_date = datetime.now(ZoneInfo("Asia/Tokyo")) + timedelta(days=-40)
    filtered_date_by_quote_result = (
        i
        for i in quote_result["data"]
        if i["department_id"] == MISUMI_TORIHIKISAKI_ID
        and datetime.fromisoformat(i["created_at"]) > from_date
    )

    # 見積一覧から必要情報を収集
    quote_list = [
        # 品目の最初の1行を使う
        QuoteData(
            durarion_src=quote["items"][0]["detail"],
            price=float(quote["subtotal_price"]),
            hinmoku_title=quote["items"][0]["name"],
        )
        for quote in filtered_date_by_quote_result
    ]

    # questionay.Choiceを使って見積書を選択する。
    # 納期（duration）を使って複数選択のデフォルト選択をマーク
    # 期日設定の 毎月26日から1か月前の日付をマーク
    # 例: 9/26締め切りの場合、8/26をマーク
    choice_list = []

    for quotedata in quote_list:
        # 対象日付の範囲を計算: 今月の26日から1か月前の日付の間か判断してデフォルト選択
        choice_list.append(
            questionary.Choice(
                title=f"{quotedata.only_katasiki} |{quotedata.price} | {quotedata.durarion}",
                value=quotedata,
                checked=is_target_date(
                    str_to_datetime_with_dateutil(quotedata.durarion),
                    datetime(today_datetime.year, today_datetime.month, 26)
                    - relativedelta(months=1),
                    datetime(today_datetime.year, today_datetime.month, 27),
                ),
            )
        )
    # 見積書の一覧を表示して、選択させる
    ask_choiced_quote_list: list[QuoteData] = questionary.checkbox(
        "請求書にする見積書を選択してください。", choices=choice_list
    ).ask()

    # 見積書の情報を元に、金額の合計を出す
    billing_data = BillingInfo(
        sum((i.price for i in ask_choiced_quote_list)),
        "ガススプリング配管図作製費",
        f"{today_datetime:%Y年%m月}請求分",
    )

    # ここで請求書情報を出して、こちらの検証と正しいか確認
    print(
        f"""
    [請求情報]
    件数: {len(ask_choiced_quote_list)}
    合計金額:{billing_data.price}
    """
    )

    # 確認用に見積書一覧作成か、請求書も生成するか、キャンセルするかの判断。
    output_select = questionary.select(
        "請求額一覧を確認する？ 全部生成する？",
        choices=[
            questionary.Choice("請求額一覧を生成する", "check"),
            questionary.Choice("請求額一覧と請求書を生成する", "output"),
        ],
    ).ask()

    match output_select:
        case "check" | "c":
            # 見積書の一覧だけ作る
            export_xlsx_path = generate_billing_list_excel(ask_choiced_quote_list)
            print(f"一覧のみ生成しました。\nファイルパス:{export_xlsx_path}")

        case "output" | "o":
            export_xlsx_path = generate_billing_list_excel(ask_choiced_quote_list)
            billing_pdf_path = generate_billing_pdf(mfcl_session, billing_data)
            print("一覧と請求書生成しました")
            print(f"一覧xlsxファイルパス:{export_xlsx_path}\n請求書pdf:{billing_pdf_path}")

            set_draft_mail([export_xlsx_path, billing_pdf_path])
            print("メールの下書きを作成しました")
        case _:
            print("キャンセルしました")


if __name__ == "__main__":
    main()
