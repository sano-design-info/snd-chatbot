import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import ClassVar

import dateutil.parser
import dateutil.tz
import openpyxl
import pandas
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from api import googleapi
from helper import load_config, rangeconvert
from helper.regexpatterns import MSM_ANKEN_NUMBER
from run_mail_action import decode_base64url

# load config
config = load_config.CONFIG
update_sheet_id = config.get("google").get("SCHEDULE_SHEET_ID")

# Excel, Gsheetのセルアドレスのパターン。"sheetname!A1"の形式でマッチする。
range_addr_pattern = re.compile(
    r"^(?P<sheetname>.*)!(?P<firstcolumn>[A-Z]+)(?P<firstrow>\d+)"
)

google_cred: Credentials = googleapi.get_cledential(googleapi.API_SCOPES)


def convert_gmail_datetimestr(gmail_datetimeformat: str) -> datetime:
    persed_time = dateutil.parser.parse(gmail_datetimeformat)
    return persed_time.astimezone(dateutil.tz.gettz("Asia/Tokyo"))


@dataclass
class ExpandedMessageItem:
    "メッセージ一覧の選択や、メール回りで使うときに利用する"
    gmail_message: dict
    payload: dict = field(init=False)
    headers: dict = field(init=False)

    id: str = field(init=False)
    title: str = field(init=False)
    subject: str = field(init=False)
    from_address: str = field(init=False)
    to_address: str = field(init=False)
    cc_address: str = field(init=False)
    datetime_: datetime = field(init=False)
    body_related: dict = field(init=False)
    body_parts: dict = field(init=False)
    body: str = field(init=False)

    def __post_init__(self):
        self.payload = self.gmail_message.get("payload")
        self.headers = self.payload.get("headers")

        self.id = self.gmail_message.get("id")
        self.title = next(
            (i for i in self.headers if i.get("name").lower() == "Subject".lower())
        ).get("value")
        self.subject = self.title
        self.from_address = next(
            (i for i in self.headers if i.get("name").lower() == "From".lower())
        ).get("value")

        # toとccは複数アドレスがあるので、",でjoinする
        self.to_address = ",".join(
            (
                i.get("value", "")
                for i in self.headers
                if i.get("name").lower() == "To".lower()
            )
        )
        self.cc_address = ",".join(
            (
                i.get("value", "")
                for i in self.headers
                if i.get("name").lower() == "Cc".lower()
            )
        )

        self.datetime_ = convert_gmail_datetimestr(
            next(
                (i for i in self.headers if i.get("name").lower() == "Date".lower())
            ).get("value")
        )

        # メールのmimeマルチパートを考慮して、構造が違うモノに対応する
        # メールがリッチテキストかつimgファイルがある場合は、multipart/relatedとなり、body_relatedを入れるとimgファイル収集も可能なので、別で用意している
        self.body_related = {}

        # ここはtext plane or multipart/altanative or multipart/related >  multipart/altanative の構造になってるらしいので、分離した処理に切り替えないといけない

        # partsがない場合 = シンプルなテキストベースの場合
        if not self.payload.get("parts"):
            self.body_parts = [self.payload]
        else:
            # リッチテキスト系の場合
            mail_part_mimetype = next(
                i.get("mimeType")
                for i in self.payload.get("parts")
                if i.get("partId") in ("0")
            )

            match mail_part_mimetype:
                case "text/plain":
                    self.body_parts = self.payload.get("parts")
                case "multipart/alternative":
                    self.body_parts = next(
                        (
                            i
                            for i in self.payload.get("parts")
                            if i.get("mimeType") == "multipart/alternative"
                        ),
                        {},
                    ).get("parts")
                case "multipart/related":
                    self.body_related = next(
                        (
                            i
                            for i in self.payload.get("parts")
                            if i.get("mimeType") == "multipart/related"
                        )
                    )
                    self.body_parts = next(
                        (
                            i
                            for i in self.body_related.get("parts")
                            if i.get("mimeType") == "multipart/alternative"
                        )
                    ).get("parts")
                case _:
                    pass

        # body_partsからbodyを取得する
        mailbody = next(
            (
                i["body"]["data"]
                for i in self.body_parts
                if "text/plain" in i.get("mimeType")
            )
        )
        self.body = decode_base64url(mailbody).decode("utf8")


@dataclass
class RenrakukoumokuInfo:
    renrakukoumoku_path: Path
    anken_base_number: str = field(init=False)
    kokyaku_name: str = field(init=False)
    enduser_name: str = field(init=False)
    format_version: str = field(init=False)

    renrakukoumoku_version_pattern: ClassVar[dict] = {
        "before202211": {"kokyaku_name_celaddr": "D6", "enduser_name_celaddr": "D9"},
        "new": {"kokyaku_name_celaddr": "D6", "enduser_name_celaddr": "D10"},
    }

    def __post_init__(self):
        # 念のために前処理でスペースがあれば除去してる
        self.anken_base_number = MSM_ANKEN_NUMBER.search(
            self.renrakukoumoku_path.name.replace(" ", "")
        ).group(1)

        renrakukoumoku_ws = openpyxl.load_workbook(self.renrakukoumoku_path).active
        version_pattern_check = renrakukoumoku_ws["B9"].value
        version_cell_addr = {}
        match version_pattern_check:
            case "エンドユーザー":
                version_cell_addr = self.renrakukoumoku_version_pattern["before202211"]
                self.format_version = "before202211"
            case "顧客連絡先":
                version_cell_addr = self.renrakukoumoku_version_pattern["new"]
                self.format_version = "new"
            case _:
                print("連絡項目の内容が不正の可能性があります")
                return None

        self.kokyaku_name = renrakukoumoku_ws[
            version_cell_addr["kokyaku_name_celaddr"]
        ].value
        self.enduser_name = renrakukoumoku_ws[
            version_cell_addr["enduser_name_celaddr"]
        ].value


@dataclass
class CsvFileInfo:
    csv_filepath: Path
    anken_number: str = field(init=False)
    anken_base_number: str = field(init=False)
    hose_type: str = field(init=False)
    hose_attachment_types: set[str] = field(init=False)

    def __post_init__(self):
        # 前処理: ファイル名にスペースとアンダーバーが入ることがあるので、置き換え
        self.anken_number = self.csv_filepath.stem.replace(" ", "").replace("_", "-")

        # 正規表現で MA-0000, MA-0000-1という
        self.anken_base_number = MSM_ANKEN_NUMBER.search(self.anken_number).group(1)

        csv_pd = pandas.read_csv(self.csv_filepath, encoding="shift-jis")
        hose_parts_pd = csv_pd[csv_pd["品名"] == "ホース(継手付)"]

        self.hose_type = hose_parts_pd["備考"].drop_duplicates().values[0]

        # 型式寸法から、SS,SL,LLを取り出して、重複を外して種類を確定させる
        self.hose_attachment_types = "/".join(
            {item.split("-")[1] for item in hose_parts_pd["型式・寸法"].values}
        )


# TODO:2023-01-12 ここではまだ一括で登録をする作業はできないので、gsheet利用優先で実装中
@dataclass
class EstimateCalcSheetInfo:
    # gsheet_url | openpyxl.ws を受け取るような仕様にする。
    calcsheet_source: str | Path
    anken_number: str = field(init=False)
    anken_base_number: str = field(init=False)
    calcsheet_parents: list[str] = field(init=False)
    duration: datetime = field(init=False)
    duration_str: str = field(init=False)
    # TODO:2023-04-19 ここはstrだが、利用する場所でintに置き換えるのでstrで良い
    price: str = field(init=False)

    def __post_init__(self):
        # sourceの種類で取り込む処理を変える
        match self.calcsheet_source:
            # excel形式: Pathを指定する
            case Path():
                if self.calcsheet_source.suffix == ".xlsx":
                    pass
                pass
            # gsheet形式: IDの羅列なのでIDが利用できるかはAPIに問い合わせる
            case str():
                # TODO:2023-04-19 ここはtryの中身が多すぎるので、API問い合わせ事にexceptする。
                try:
                    sheet_service = build("sheets", "v4", credentials=google_cred)

                    # スプレッドシート名を収集して、anken_numberを生成

                    self.gsheet_values = (
                        sheet_service.spreadsheets()
                        .get(
                            spreadsheetId=self.calcsheet_source,
                        )
                        .execute()
                    )

                    self.anken_number = MSM_ANKEN_NUMBER.search(
                        self.gsheet_values["properties"]["title"]
                    ).group(0)
                    self.anken_base_number = MSM_ANKEN_NUMBER.search(
                        self.anken_number
                    ).group(1)

                    # 古い仕様のシートと新しいシートの違いで、range_mapを変える。計算結果シートがある場合はそこを参照
                    estimate_calc_sheetnames = [
                        sheets["properties"]["title"]
                        for sheets in self.gsheet_values["sheets"]
                    ]
                    # print(estimate_calc_sheetnames)
                    # TODO:2023-01-12 この部分は判断方法と結果を返す関数にした方がいいと思われる。クラスの裏に分離したほうがいいな
                    range_map = {
                        "Sheet1!F17": "price",
                        "Sheet1!F1": "duration",
                    }

                    if "計算結果" in estimate_calc_sheetnames:
                        range_map = {
                            "'計算結果'!B5": "price",
                            "'計算結果'!B6": "duration",
                        }

                    # シートの情報を収集して各フィールドへ追加す各
                    range_names = list(range_map.keys())
                    estimate_calc_gsheet_values_res = (
                        sheet_service.spreadsheets()
                        .values()
                        .batchGet(
                            spreadsheetId=self.calcsheet_source,
                            ranges=range_names,
                            # valueRenderOption="UNFORMATTED_VALUE",
                        )
                        .execute()
                    )
                    for res_value in estimate_calc_gsheet_values_res.get("valueRanges"):
                        setattr(
                            self,
                            range_map[res_value.get("range")],
                            str(res_value.get("values")[0][0]),
                        )
                    # gsheetで取り込んだ結果が数字になってしまう...
                    self.duration = self.fix_datetime(self.duration)
                    self.duration_str = self.duration.strftime("%m/%d")
                    self.price = re.sub(r"[\¥\,]", "", self.price)

                except HttpError as error:
                    # TODO:2022-12-09 エラーハンドリングは基本行わずここで落とすこと
                    # TODO: 2022/12/28 このエラーは致命的なのでそのままプログラム自体も終了する
                    print(f"An error occurred: {error}")
                    exit()
            # Path, str以外はエラーとする
            case _:
                raise ValueError(
                    f"This source is cant use class:{self.calcsheet_source}"
                )

    def fix_datetime(datetime_str: str) -> datetime:
        """
        日付の入力の区切り文字を修正する。
        """
        for splitecahr in (".", "/"):
            if splitecahr in datetime_str:
                return datetime.strptime(
                    datetime_str, splitecahr.join(("%Y", "%m", "%d"))
                )


# TODO:2022-12-29 案件マップというよりは、コンバーターという名前のほうがしっくりくる
@dataclass
class MsmAnkenMap:
    """
    各案件の情報をまとめるためのクラス
    anken_number, anken_basenumberのフォーマットは "MA-0000" のような"MA-"の接頭辞ありのものを想定している
    """

    csvfile_info: CsvFileInfo | None = None
    estimate_calcsheet_info: EstimateCalcSheetInfo | None = None
    anken_number: str = field(init=False)
    anken_base_number: str = field(init=False)
    # 連絡項目と関連するものを入れるフィールド
    renrakukoumoku_info: RenrakukoumokuInfo | None = None

    def __post_init__(self):
        # TODO:2022-12-29 今はcsvfile_infoがあることが前提だが、estimate_sheet_infoのどちらか、または両方ある場合の対応を作ること
        self.anken_number = self.select_anken_ref().anken_number
        self.anken_base_number = self.select_anken_ref().anken_base_number

    def select_anken_ref(self):
        anken_ref_obj = None
        if self.csvfile_info:
            anken_ref_obj = self.csvfile_info
        elif self.estimate_calcsheet_info:
            anken_ref_obj = self.estimate_calcsheet_info
        return anken_ref_obj

    # 書くデータのクラスを取り込む関数（set_***）
    # setしたときにanken_base_numberが同じかをチェック
    def set_renrakukoumoku_info(self, renrakukoumoku_info: RenrakukoumokuInfo):
        if (
            renrakukoumoku_info.anken_base_number
            == self.select_anken_ref().anken_base_number
        ):
            self.renrakukoumoku_info = renrakukoumoku_info


@dataclass
class MsmAnkenMapList:
    msmankenmap_list: list[MsmAnkenMap] = field(default_factory=list)

    # 名寄せ用のマップ
    schedule_sheet_map: ClassVar[dict] = {
        "納期": "estimate_calcsheet_info.duration_str",
        "金額(税抜)": "estimate_calcsheet_info.price",
        "ガス本数": "csvfile_info.gas_qty",
        "ホース本数": "csvfile_info.hose_qty",
        "ホースタイプ": "csvfile_info.hose_type",
        "利用したホースの接続継手の種類": "csvfile_info.hose_attachment_types",
        "顧客名": "renrakukoumoku_info.kokyaku_name",
        "エンドユーザー": "renrakukoumoku_info.enduser_name",
    }

    # 更新データの生成
    # CSVor estimatesheetがない場合は処理を行わない（連絡項目だけだと正確な案件番号が出ないので）
    def generate_update_sheet_values(self) -> pandas.DataFrame:
        # 入っているmsmankenmapにcsvファイル/見積書計算のデータクラスがない場合は無視する csvfile_info.anken_numberがない場合は無視
        sheet_table = []
        for msmankeninfo in self.msmankenmap_list:
            if not msmankeninfo.anken_number:
                print("案件番号がありません。CSVファイルか計算表を登録してください")
                continue

            # 各種のデータクラスから値を取り出していく
            # TODO:2022-12-29 各データクラスのどれかがない場合の考慮がまだできてない。
            # ない場合は空白とかでいいと思う

            sheet_row = {"msmankennumber": msmankeninfo.anken_number}

            # dataclass_attr.split(".")[0]):CSV, 連絡項目などの収集infoのオブジェクト
            # dataclass_attr.split(".")[1]):それぞれ収集infoオブジェクトのインスタンス属性
            sheet_row = sheet_row | {
                sheet_column_name: getattr(
                    getattr(msmankeninfo, map_dataclass_attr.split(".")[0]),
                    map_dataclass_attr.split(".")[1],
                    None,
                )
                # sheet_column_name: map_dataclass_attr
                for sheet_column_name, map_dataclass_attr in MsmAnkenMapList.schedule_sheet_map.items()
            }
            sheet_table.append(sheet_row)

        # df変換
        return pandas.DataFrame(sheet_table).set_index(["msmankennumber"]).fillna("")


def get_schedule_table_area(
    search_range: str, google_cred
) -> tuple[str, pandas.DataFrame]:
    # 更新対称のセル範囲から値を取得
    try:
        append_values = [
            [],
        ]
        # appendを使ってテーブルの範囲を取得最終行からテーブルの範囲を取得
        sheet_service = build("sheets", "v4", credentials=google_cred)
        append_end_row = (
            sheet_service.spreadsheets()
            .values()
            .append(
                spreadsheetId=update_sheet_id,
                range=search_range,
                body={"values": append_values},
                valueInputOption="USER_ENTERED",
                insertDataOption="INSERT_ROWS",
            )
        ).execute()

        last_low_range = range_addr_pattern.match(
            append_end_row.get("updates").get("updatedRange"),
        ).group("firstrow")

        schedule_range_matcher = range_addr_pattern.match(search_range)

        schedule_table_range = f"{schedule_range_matcher.group(0)}:Q{last_low_range}"

        schedule_table_res = (
            sheet_service.spreadsheets()
            .values()
            .get(spreadsheetId=update_sheet_id, range=schedule_table_range)
            .execute()
        )
        # 値が取得できた。
        schedule_table = schedule_table_res.get("values")

        # df変換
        result_pd = pandas.DataFrame(schedule_table)
        result_pd = result_pd.set_axis(labels=schedule_table[0], axis=1).drop(
            result_pd.index[[0]]
        )
        result_pd = result_pd.set_index(["図番"], drop=False)

        return result_pd

    except HttpError as error:
        # このエラーは致命的なのでそのままプログラム自体も終了する
        print(f"An error occurred: {error}")
        exit()


def generate_update_valueranges(
    search_range: str, old_pd: pandas.DataFrame, new_pd: pandas.DataFrame
) -> list[dict]:
    """
    oldとnewの表をみて、old側で更新するべきセルのアドレスと値を収集する
    結果はGoogle Sheet APIで受け取れるValueRangeとする
    # https://developers.google.com/sheets/api/reference/rest/v4/spreadsheets.values#ValueRange
    search_range ="'sheet1'!A5:Q5"
    [{"rangeaddr": "value"}, ...]
    """

    # TODO: 2023/01/16 今回はcolumnについては何も考慮していない。A1表記のAを変換して数字にした上で、次のテーブルのループで対象のセルアドレスを生成する必要あり。今回はA列から取ってるからやらなくていいけど、実装の余裕ができたらやろう
    schedule_range_matcher = range_addr_pattern.match(search_range)

    start_row = int(schedule_range_matcher.group("firstrow")) + 1

    result_rangevalues = []
    new_pd_columns = new_pd.columns.tolist()
    new_pd_indexs = new_pd.index.tolist()

    for table_row_index, (old_index, row) in enumerate(old_pd.iterrows()):
        for table_column_index, (old_column_name, column_by_row) in enumerate(
            row.items()
        ):
            # oldのインデックスとカラム名がnew側にない場合は無視
            if not (old_index in new_pd_indexs and old_column_name in new_pd_columns):
                continue
            # 上書き防止条件: newに値がある, old側に値がない場合は追加する
            if new_pd.loc[old_index, old_column_name] and not column_by_row:
                cell_addr = rangeconvert.rowcol_to_a1(
                    start_row + table_row_index, table_column_index + 1
                )
                # 上のrangeaddr, valueの辞書を作る。
                result_rangevalues.append(
                    {
                        "range": cell_addr,
                        "values": [[new_pd.loc[old_index, old_column_name]]],
                    }
                )
    return result_rangevalues


def update_schedule_sheet(update_data: list[dict], google_cred):
    # 最後にbatchUpdateを行うためのデータ構造を生成して流し込む
    update_value_body = {
        "data": [update_data],
        "valueInputOption": "USER_ENTERED",
    }

    sheet_service = build("sheets", "v4", credentials=google_cred)
    try:
        update_gsheet_res = (
            sheet_service.spreadsheets()
            .values()
            .batchUpdate(spreadsheetId=update_sheet_id, body=update_value_body)
            .execute()
        )
        return update_gsheet_res
    except HttpError as error:
        # TODO:2022-12-09 エラーハンドリングは基本行わずここで落とすこと
        print(f"An error occurred: {error}")
        exit()
